from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from .forms import ConductorForm, ReporteForm
from .models import conductores
from .models import rutas, infracciones, ReporteProgramacion, vehiculos, reporteGeneral
from tablib import Dataset
from .resources import conductoresResource
from openpyxl import load_workbook


def home(request):
    return render(request, 'home.html')


def signup(request):
    if request.method == 'GET':
        return render(request, 'signup.html', {
            'form': UserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(username=request.POST['username'],
                                                password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('horarios')
            except IntegrityError:
                return render(request, 'signup.html', {
                    'form': UserCreationForm,
                    'error': 'El usuario ya existe!'
                })
        return render(request, 'signup.html', {
            'form': UserCreationForm,
            'error': 'Las contraseñas no coinciden!'
        })


def horarios(request):
    return render(request, 'horarios.html')


def crear_conductor(request):
    if request.method == 'GET':
        return render(request, 'crear_conductor.html', {
            'form': ConductorForm
        })
    else:
        try:
            form = ConductorForm(request.POST)
            nuevo_conductor = form.save(commit=False)
            nuevo_conductor.user = request.user
            nuevo_conductor.save()
            return redirect('horarios')
        except ValueError:
            return render(request, 'crear_conductor.html'), {
                'form': ConductorForm,
                'error': 'No se pudo crear el registro'
            }


def ver_conductores(request):
    listado = conductores.objects.all()
    return render(request, 'ver_conductores.html', {'conductores': listado})


def ver_vehiculos(request):
    listado = vehiculos.objects.all()
    return render(request, 'ver_vehiculos.html', {'vehiculos': listado})


def ver_detalle_conductor(request, id):
    ver_conductor = get_object_or_404(conductores, pk=id)
    return render(request, 'ver_detalle_conductor.html', {'conductores': ver_conductor})


def signout(request):
    logout(request)
    return redirect('home')


def sigin(request):
    if request.method == 'GET':
        return render(request, 'sigin.html', {
            'form': AuthenticationForm
        })
    else:
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'sigin.html', {
                'form': AuthenticationForm,
                'error': 'El usuario o la contraseña es incorrecto'
            })
        else:
            login(request, user)
            return redirect('horarios')


def ver_rutas(request):
    listado_rutas = rutas.objects.all()
    return render(request, 'rutas.html', {'rutas': listado_rutas})


def ver_infracciones(request):
    listado_infracciones = infracciones.objects.all()
    return render(request, 'infracciones.html', {'infracciones': listado_infracciones})


def reportes(request):
    listado_reportes = reporteGeneral.objects.all()
    return render(request, 'reportes.html', {'reportes': listado_reportes})


def crear_reporte(request):
    if request.method == 'POST':
        return render(request, 'crear_reporte.html', {
            'form': ReporteForm
        })
    else:
        try:
            form = ReporteForm(request.POST)
            nuevo_reporte = form.save(commit=False)
            # nuevo_reporte.user = request.user
            nuevo_reporte.save()
            return redirect('reportes')
        except ValueError:
            return render(request, 'crear_reporte.html'), {
                'form': ReporteForm,
                'error': 'No se pudo crear el registro'
            }


def importar_conductores(request):
    if request.method == 'POST' and request.FILES['archivo_excel']:
        archivo = request.FILES['archivo_excel']
        # Cargamos el archivo Excel utilizando openpyxl
        libro = load_workbook(archivo)
        hoja = libro.active

        for fila in hoja.iter_rows(min_row=0, values_only=True):
            # Aquí asumimos que la estructura de la hoja de Excel coincide con los campos del modelo
            documento = fila[0]
            nombre_completo = fila[1]
            tipo_licencia = fila[2]
            fecha_nacimiento = fila[3]
            telefono = fila[4]
            bus_asignado = fila[5]
            alerta = fila[6]

            # Creamos un objeto de Conductores y lo guardamos en la base de datos
            conductor = conductores(
                documento=documento,
                nombre_completo=nombre_completo,
                tipo_licencia=tipo_licencia,
                fechaNacimiento=fecha_nacimiento,
                telefono=telefono,
                bus_asignado=bus_asignado,
                alerta=alerta
            )
            conductor.save()

        return render(request, 'importacion_exitosa.html')

    return render(request, 'importar_conductores.html')


def importar_vehiculos(request):
    if request.method == 'POST' and request.FILES['archivo_excel']:
        archivo = request.FILES['archivo_excel']
        # Cargamos el archivo Excel utilizando openpyxl
        libro = load_workbook(archivo)
        hoja = libro.active

        for fila in hoja.iter_rows(min_row=0, values_only=True):
            # Aquí asumimos que la estructura de la hoja de Excel coincide con los campos del modelo
            numero_interno = fila[0]
            tipo = fila[1]
            placa = fila[2]
            documento = fila[3]

            # Creamos un objeto de Conductores y lo guardamos en la base de datos
            vehiculo = vehiculos(
                numero_interno=numero_interno,
                tipo=tipo,
                placa=placa,
                documento=documento,

            )
            vehiculo.save()

        return render(request, 'importacion_exitosa.html')

    return render(request, 'importar_vehiculos.html')


def importar_reportes(request):
    if request.method == 'POST' and request.FILES['archivo_excel']:
        archivo = request.FILES['archivo_excel']
        # Cargamos el archivo Excel utilizando openpyxl
        libro = load_workbook(archivo)
        hoja = libro.active

        for fila in hoja.iter_rows(min_row=0, values_only=True):
            # Aquí asumimos que la estructura de la hoja de Excel coincide con los campos del modelo
            documento = fila[0]
            nombre_completo = fila[1]
            bus_asignado = fila[2]
            tipo_bus_asignado = fila[3]
            placa_bus_asignado = fila[4]
            ruta_asignada = fila[5]
            viajes_programados = fila[6]
            tabla_asignada = fila[7]


            # Creamos un objeto de Conductores y lo guardamos en la base de datos
            reporte_general = reporteGeneral(
                documento=documento,
                nombre_completo=nombre_completo,
                bus_asignado=bus_asignado,
                tipo_bus_asignado=tipo_bus_asignado,
                placa_bus_asignado=placa_bus_asignado,
                ruta_asignada=ruta_asignada,
                viajes_programados=viajes_programados,
                tabla_asignada=tabla_asignada,

            )
            reporte_general.save()

        return render(request, 'importacion_exitosa.html')

    return render(request, 'importar_reportes.html')